"""
Excel -> Supabase Timetable Importer
=====================================
Reads the college Excel timetable (UG, PG TIME TABLE JAN TO MAY 2026.xlsx)
and generates Python subgroup files OR inserts directly into Supabase.

Usage:
    # Preview what will be parsed (dry run, no files written)
    python timetable_importer/import_from_excel.py --sheet "1ST YEAR A" --year 1 --dry-run

    # Generate Python files for all subgroups in a sheet
    python timetable_importer/import_from_excel.py --sheet "1ST YEAR A" --year 1

    # Import specific subgroups only
    python timetable_importer/import_from_excel.py --sheet "1ST YEAR A" --year 1 --subgroups 1A11 1A12

    # Push directly to Supabase (requires DATABASE_URL env var)
    python timetable_importer/import_from_excel.py --sheet "1ST YEAR A" --year 1 --push-to-db
"""

import os
import re
import sys
import argparse
import datetime
import openpyxl
from pathlib import Path

# ---- Paths ------------------------------------------------------------------
BASE_DIR   = Path(__file__).parent.parent
EXCEL_PATH = BASE_DIR / "docs" / "resources" / "UG, PG TIME TABLE JAN TO MAY 2026.xlsx"
OUTPUT_DIR = BASE_DIR / "subgroups"

# ---- Subject name lookup (code -> full name) ---------------------------------
# Add more codes here as you encounter them for new departments/years
SUBJECT_NAMES = {
    # 1st Year Common
    "UPH013": "Engineering Physics",
    "UES101": "Engineering Chemistry",
    "UES102": "Engineering Drawing and Design",
    "UMA023": "Engineering Mathematics I",
    "UHU003": "Communicative English",
    "UMA024": "Engineering Mathematics II",
    # 1st Year B
    "UCB009": "Introduction to Computing",
    "UES013": "Engineering Workshop",
    "UMA022": "Engineering Mathematics",
    "UES103": "Engineering Mechanics",
    # 2nd Year CSE
    "UCS301": "Data Structures",
    "UCS302": "Digital Electronics",
    "UCS303": "Computer Organization",
    "UCS304": "Discrete Mathematics",
    "UCS305": "Object Oriented Programming",
    # Add more codes below as needed:
}

# ---- Day parsing ------------------------------------------------------------
DAY_MAP = {
    'M': 0,  # Monday
    'T': 1,  # Tuesday
    'W': 2,  # Wednesday
    'H': 3,  # Thursday (some sheets use H)
    'F': 4,  # Friday
    'S': 5,  # Saturday
}

# ---- Type suffix mapping ----------------------------------------------------
def get_type(code_str):
    """
    Split 'UPH013P' into ('UPH013', 'Lab', slots).
    Suffix: P -> Lab (2 slots / 100min), T -> Tutorial (1 slot), L/none -> Class (1 slot)
    Returns ('', '', 0) for non-subject codes (e.g. 'APC-1', 'NSD', room names).
    """
    code_str = str(code_str).strip()
    if not code_str:
        return ('', '', 1)
    m = re.match(r'^([A-Z]{2,4}\d{3})([A-Z]?)$', code_str)
    if m:
        code   = m.group(1)
        suffix = m.group(2).upper()
        if suffix == 'P':
            return (code, 'Lab', 2)       # 2 slots = 100 min
        elif suffix == 'T':
            return (code, 'Tutorial', 1)
        elif suffix == 'L':
            return (code, 'Class', 1)
        else:
            return (code, 'Class', 1)
    return ('', '', 1)   # not a subject code


def parse_time(val):
    """Normalise various time formats to 'HH:MM'."""
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val.strftime('%H:%M')
    s = str(val).strip().upper()
    s = s.replace(' ', '').replace(':AM', '').replace(':PM', '')
    parts = s.split(':')
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        return f"{h:02d}:{m:02d}"
    except Exception:
        return None


def add_50min(t):
    """Add 50 minutes to a 'HH:MM' time string."""
    h, m = map(int, t.split(':'))
    total = h * 60 + m + 50
    return f"{total // 60:02d}:{total % 60:02d}"


# ---- Core parser ------------------------------------------------------------
def parse_sheet(ws):
    """
    Parse one Excel sheet into a dict:
        { subgroup_code: { 'subjects': set(codes), 'timetable': [(code,day,start,end,type)] } }
    """
    rows = list(ws.iter_rows(values_only=True))

    # Step 1: find header row (contains 'DAY' and 'HOURS')
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip().upper() if c else '' for c in row]
        if 'DAY' in cells and 'HOURS' in cells:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find header row (DAY / HOURS) in sheet")

    header_row = rows[header_idx]

    hours_col = next(
        i for i, c in enumerate(header_row)
        if str(c).strip().upper() == 'HOURS'
    )
    day_col = 0

    # Step 2: find subgroup columns (0-3 rows above/at header)
    # look_back=0 checks the header row itself (practical codes like 1A1A)
    # look_back=1 checks one row above (tutorial codes like 1A11)
    subgroup_cols = {}   # col_index -> subgroup_code
    for look_back in range(0, 4):
        candidate_idx = header_idx - look_back
        if candidate_idx < 0:
            continue
        cand = rows[candidate_idx]
        for col_i, val in enumerate(cand):
            if col_i <= hours_col:
                continue
            sv = str(val).strip() if val else ''
            if re.match(r'^\d[A-Z]\w{1,3}$', sv):
                if col_i not in subgroup_cols:
                    subgroup_cols[col_i] = sv

    if not subgroup_cols:
        raise ValueError("Could not detect any subgroup columns")

    # Step 3: build slot-number -> start_time map
    slot_times = {}
    srnr_col = next(
        (i for i, c in enumerate(header_row)
         if str(c).strip().upper() in ('SR NO', 'SR.NO', 'SRNO')),
        1
    )
    for row in rows[header_idx + 1:]:
        sr_val = row[srnr_col] if len(row) > srnr_col else None
        hr_val = row[hours_col] if len(row) > hours_col else None
        if sr_val and isinstance(sr_val, (int, float)):
            t = parse_time(hr_val)
            if t:
                slot_times[int(sr_val)] = t

    # Step 4: parse data rows
    result = {sg: {'subjects': set(), 'timetable': []} for sg in subgroup_cols.values()}

    current_day   = None
    current_slot  = None
    current_start = None
    is_room_row   = False

    for row in rows[header_idx + 1:]:
        if len(row) == 0:
            continue

        day_cell = str(row[day_col]).strip().upper() if row[day_col] else ''
        if day_cell in DAY_MAP:
            current_day = DAY_MAP[day_cell]
            is_room_row = False

        sr_val = row[srnr_col] if len(row) > srnr_col else None
        if sr_val and isinstance(sr_val, (int, float)):
            current_slot  = int(sr_val)
            current_start = slot_times.get(current_slot)
            is_room_row   = False

        if current_day is None or current_start is None:
            is_room_row = not is_room_row
            continue

        if is_room_row:
            is_room_row = False
            continue

        # Subject row
        for col_i, sg in subgroup_cols.items():
            if col_i >= len(row):
                continue
            cell_val = str(row[col_i]).strip() if row[col_i] else ''
            skip_vals = {'lab', 'lab1', 'lab-1', 'lab2', 'lab-2', 'lkb', 'lkb1', 'lab3', 'lab-3'}
            if not cell_val or cell_val.lower() in skip_vals:
                continue

            code, cls_type, slots = get_type(cell_val)
            if not code:
                continue

            # Calculate end time: 50min per slot
            if slots == 2:
                end_time = add_50min(add_50min(current_start))  # Lab = 100min
            else:
                end_time = add_50min(current_start)             # Tutorial/Class = 50min

            # Merge adjacent same-subject entries (safety net if same code appears again)
            entries = result[sg]['timetable']
            if entries:
                last = entries[-1]
                if last[0] == code and last[1] == current_day and last[3] == current_start:
                    entries[-1] = (code, current_day, last[2], end_time, cls_type)
                    result[sg]['subjects'].add(code)
                    continue

            entries.append((code, current_day, current_start, end_time, cls_type))
            result[sg]['subjects'].add(code)

        is_room_row = True

    return result


# ---- Python file generator --------------------------------------------------
def generate_py_file(subgroup, data, year):
    subjects  = sorted(data['subjects'])
    timetable = sorted(data['timetable'], key=lambda x: (x[1], x[2]))

    subj_lines = []
    for code in subjects:
        name = SUBJECT_NAMES.get(code, f"Subject {code}")
        subj_lines.append(f'    ("{name}", "{code}"),')

    tt_lines = []
    for entry in timetable:
        code, day, start, end, cls_type = entry[:5]
        tt_lines.append(f'    ("{code}", {day}, "{start}", "{end}", "{cls_type}"),')

    lines = [
        f'SUBGROUP = "{subgroup}"',
        f'YEAR     = {year}',
        '',
        'subjects = [',
        *subj_lines,
        ']',
        '',
        'timetable = [',
        *tt_lines,
        ']',
        '',
    ]
    return '\n'.join(lines)


# ---- Direct DB push ---------------------------------------------------------
def push_to_db(all_data, year):
    sys.path.insert(0, str(BASE_DIR))
    from db import get_db_connection, execute_query

    conn = get_db_connection()
    total = 0
    for subgroup, data in all_data.items():
        if not data['timetable']:
            continue
        try:
            subj_ids = {}
            for code in data['subjects']:
                name = SUBJECT_NAMES.get(code, f"Subject {code}")
                execute_query(conn,
                    """INSERT INTO subjects (name, code, subgroup, year)
                       VALUES (%s, %s, %s, %s)
                       ON CONFLICT (code, subgroup) DO UPDATE SET name = EXCLUDED.name""",
                    (name, code, subgroup, year))
                res = execute_query(conn,
                    "SELECT id FROM subjects WHERE code = %s AND subgroup = %s",
                    (code, subgroup))
                subj_ids[code] = res[0]['id'] if res else None

            for entry in data['timetable']:
                code, day, start, end, cls_type = entry[:5]
                weight = entry[5] if len(entry) > 5 else None
                subj_id = subj_ids.get(code)
                if not subj_id:
                    continue
                execute_query(conn,
                    """INSERT INTO timetable
                       (subject_id, subgroup, day_of_week, start_time, end_time, type, weight_override)
                       VALUES (%s, %s, %s, %s, %s, %s, %s)
                       ON CONFLICT DO NOTHING""",
                    (subj_id, subgroup, day, start, end, cls_type, weight))

            conn.commit()
            print(f"  [OK] {subgroup}: {len(data['subjects'])} subjects, {len(data['timetable'])} slots")
            total += 1
        except Exception as e:
            conn.rollback()
            print(f"  [FAIL] {subgroup}: {e}")

    conn.close()
    print(f"\nDone: {total} subgroups pushed to database.")


# ---- CLI entry --------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Import timetable from Excel into Attendex')
    parser.add_argument('--sheet',      required=True, help='Sheet name, e.g. "1ST YEAR A"')
    parser.add_argument('--year',       type=int, default=1, help='Academic year (1-4)')
    parser.add_argument('--subgroups',  nargs='*', help='Only process these subgroup codes')
    parser.add_argument('--dry-run',    action='store_true', help='Preview without writing files')
    parser.add_argument('--push-to-db', action='store_true', help='Insert into Supabase directly')
    parser.add_argument('--excel',      default=str(EXCEL_PATH), help='Path to Excel file')
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    print(f"Loading: {excel_path.name}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if args.sheet not in wb.sheetnames:
        print(f"ERROR: Sheet '{args.sheet}' not found.")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[args.sheet]
    print(f"Parsing sheet: {args.sheet}\n")

    try:
        all_data = parse_sheet(ws)
    except ValueError as e:
        print(f"Parse error: {e}")
        sys.exit(1)

    if args.subgroups:
        all_data = {k: v for k, v in all_data.items() if k in args.subgroups}

    print(f"Found {len(all_data)} subgroups:\n")
    for sg, data in all_data.items():
        n_subj  = len(data['subjects'])
        n_slots = len(data['timetable'])
        print(f"  {sg}: {n_subj} subjects, {n_slots} timetable slots")
        if args.dry_run:
            for entry in data['timetable'][:5]:
                print(f"      {entry}")
            if len(data['timetable']) > 5:
                print(f"      ... and {len(data['timetable'])-5} more")

    if args.dry_run:
        print("\n[Dry run - no files written]")
        return

    if args.push_to_db:
        print("\nPushing to database...")
        push_to_db(all_data, args.year)
        return

    # Write Python files
    print(f"\nWriting Python files to subgroups/...")
    for sg, data in all_data.items():
        if not data['timetable']:
            print(f"  [SKIP] {sg}: no timetable data")
            continue

        folder_name = sg[:2].upper()
        folder = OUTPUT_DIR / folder_name
        folder.mkdir(parents=True, exist_ok=True)

        out_path = folder / f"TT_{sg}.py"
        content  = generate_py_file(sg, data, args.year)

        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(content)

        print(f"  [OK] Written: {out_path.relative_to(BASE_DIR)} ({len(data['timetable'])} slots)")

    print(f"\nDone! To push to DB, run:")
    print(f"  python subgroups/seed_all.py --force")


if __name__ == '__main__':
    main()
